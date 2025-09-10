#********************************************************************
# Nom : AlgorythmExportExcelToJson.py
# Par : Corentin Couëron
# Date : 09/09/2025
# Version : 1.0
# Ce programme exporte des donnees d'un fichier Excel vers un fichier JSON.
# Algorithme supplémentaire : 
# - Elimination des doublons
# - Tri des mots par ordre alphabetique
#*********************************************************************

# Librairies
import json
from openpyxl import load_workbook

# Chargement du fichier Excel
wb = load_workbook("E:\P003 - Word Hunt\AlgorythmExportExcelToJson\LexiqueFr.xlsx", data_only=True)
ws = wb.active

# Ensemble pour stocker les mots uniques
vus = set()

# Collecte unique
for row in ws.iter_rows(values_only=True):
    valeur = row[2]  # 3e colonne du fichier Excel
    if valeur is not None:
        mot = str(valeur).strip()
        vus.add(mot)

# Tri alphabétique
resultat = [
    {"word": mot, "found": False}
    for mot in sorted(vus, key=str.lower)  # trie insensible à la casse
]

# Sauvegarde JSON
with open("LexiqueFr.json", "w", encoding="utf-8") as f:
    json.dump(resultat, f, indent=2, ensure_ascii=False)

print(f"Export terminé ! {len(resultat)} mots uniques triés.")
